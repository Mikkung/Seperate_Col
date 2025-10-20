import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

st.set_page_config(page_title="แยกชีตตามอาจารย์ผู้สอน", page_icon="📘")

st.title("📘 แยกชีตในไฟล์ Excel ตามอาจารย์ผู้สอน (พร้อมสรุปและ merge cell)")

uploaded_file = st.file_uploader("📤 อัปโหลดไฟล์ Excel (.xlsx)", type=["xlsx"])

if uploaded_file:
    df = pd.read_excel(uploaded_file)

    # ✅ ถ้า column แรกไม่มีชื่อ -> ตั้งชื่อให้ว่า "ลำดับ"
    if not str(df.columns[0]).strip() or "unnamed" in str(df.columns[0]).lower():
        df.columns = ["ลำดับ"] + list(df.columns[1:])
    else:
        df.rename(columns={df.columns[0]: "ลำดับ"}, inplace=True)

    # หาชื่อคอลัมน์ "อาจารย์ผู้สอน"
    teacher_col = next((c for c in df.columns if "อาจารย์" in str(c)), None)

    if not teacher_col:
        st.error("❌ ไม่พบคอลัมน์ชื่อ 'อาจารย์ผู้สอน'")
    else:
        st.success(f"✅ พบคอลัมน์อาจารย์ผู้สอน: '{teacher_col}'")

        if st.button("🚀 สร้างไฟล์แยกชีตพร้อมสรุป"):
            output = BytesIO()

            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                for teacher, group in df.groupby(teacher_col):
                    group = group.copy()

                    # 🧹 ตัดคอลัมน์สุดท้ายออก
                    group = group.iloc[:, :-1]

                    # ✅ ตั้งชื่อคอลัมน์แรกให้แน่ใจว่าเป็น "ลำดับ"
                    group.rename(columns={group.columns[0]: "ลำดับ"}, inplace=True)

                    # รีเซ็ตลำดับใหม่
                    group["ลำดับ"] = range(1, len(group) + 1)

                    # หาคอลัมน์จำนวนนิสิตและจำนวนเงิน
                    student_col = next((c for c in group.columns if "นิสิต" in str(c)), None)
                    money_col = next((c for c in group.columns if "เงิน" in str(c)), None)

                    # แปลงข้อมูลให้เป็นตัวเลข
                    if student_col:
                        group[student_col] = pd.to_numeric(group[student_col], errors='coerce').fillna(0)
                    if money_col:
                        group[money_col] = group[money_col].astype(str).str.replace(',', '', regex=False)
                        group[money_col] = pd.to_numeric(group[money_col], errors='coerce').fillna(0)

                    # รวมค่า
                    total_students = group[student_col].sum() if student_col else ""
                    total_money = group[money_col].sum() if money_col else ""

                    # เพิ่มแถวสรุป
                    summary = pd.DataFrame({
                        "ลำดับ": ["รวมเป็นเงิน"],
                        student_col: [total_students],
                        money_col: [total_money]
                    })
                    for col in group.columns:
                        if col not in summary.columns:
                            summary[col] = ""

                    summary = summary[group.columns]
                    final_df = pd.concat([group, summary], ignore_index=True)

                    # เขียนลงชีต
                    safe_name = str(teacher).strip()[:31].replace('/', '-')
                    final_df.to_excel(writer, sheet_name=safe_name, index=False)

            # โหลด workbook เพื่อ merge cell + style
            output.seek(0)
            wb = load_workbook(output)

            for ws in wb.worksheets:
                last_row = ws.max_row
                last_col = ws.max_column

                # ค้นหาคอลัมน์ “จำนวนนิสิต”
                headers = [cell.value for cell in ws[1]]
                student_col_idx = headers.index("จำนวนนิสิต") + 1 if "จำนวนนิสิต" in headers else 6

                # ✅ Merge cell แถวสุดท้าย จาก col 1 ถึง ก่อน col จำนวนนิสิต
                ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=student_col_idx - 1)
                ws.cell(row=last_row, column=1).alignment = Alignment(horizontal="center", vertical="center")

                # ตั้งค่า font ให้แถวสุดท้ายเป็น bold
                for col in range(1, last_col + 1):
                    cell = ws.cell(row=last_row, column=col)
                    cell.font = Font(bold=True)

            new_output = BytesIO()
            wb.save(new_output)
            new_output.seek(0)

            st.download_button(
                label="📥 ดาวน์โหลดไฟล์ที่แยกแล้ว (พร้อมรวมเซลล์)",
                data=new_output,
                file_name="แยกตามอาจารย์_พร้อมสรุป_merge.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
